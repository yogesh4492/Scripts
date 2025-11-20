import os
import typer
from openpyxl import load_workbook

app = typer.Typer(help="Extract 'Path/File Name' from all XLSX files into a single TXT file.")

@app.command()
def extract(
    folder: str = typer.Argument(..., help="Folder containing XLSX files"),
    output_file: str = typer.Option("all_paths.txt", help="Output TXT file containing all paths")
):
    all_paths = []

    for root, _, files in os.walk(folder):
        for file in files:
            if file.lower().endswith(".xlsx"):
                file_path = os.path.join(root, file)

                try:
                    wb = load_workbook(file_path, read_only=True)
                    sheet = wb.active

                    # Read header row
                    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

                    # Required columns
                    required = ["Path", "File Name"]
                    missing = [c for c in required if c not in headers]
                    if missing:
                        typer.echo(f"⚠️ Missing columns {missing} in {file_path}")
                        continue

                    path_col = headers.index("Path")
                    fname_col = headers.index("File Name")

                    # Extract paths
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        path_val = row[path_col]
                        fname_val = row[fname_col]

                        if path_val and fname_val:
                            final_path = f"{path_val.rstrip('/')}/{fname_val.lstrip('/')}"
                            all_paths.append(final_path)

                    wb.close()
                    typer.echo(f"✔️ Processed: {file_path}")

                except Exception as e:
                    typer.echo(f"❌ Error reading {file_path}: {e}")

    # Write all paths to single TXT
    with open(output_file, "w", encoding="utf-8") as f:
        for path in all_paths:
            f.write(path + "\n")

    typer.echo(f"✅ Done! All paths saved in {output_file}")


if __name__ == "__main__":
    app()
